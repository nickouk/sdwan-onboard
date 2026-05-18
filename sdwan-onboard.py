#!/usr/bin/env python3
"""
vManage SD-WAN WAN Edge Deployment Script
Target: vManage 20.15.4.4
URL: https://vmanage-953677893.sdwan.cisco.com/

Device lookup strategy
----------------------
Devices are matched from the vManage WAN edge inventory using the full
Device ID from CSV column A (e.g. 'C1121X-8P-FCZ2937R0T3'), which maps
directly to the 'uuid' field in the vManage vedge inventory API.

The serial number suffix (after the last '-') is retained for display and
as a last-resort fallback only.

  CSV Device ID  : C1121X-8P-FCZ2937R0T3
  vManage uuid   : C1121X-8P-FCZ2937R0T3  <- direct match

Tasks
-----
  0. Read site data from CSV (column D = Site Id)
  1. Connect to vManage; locate devices by serial number
  2. Check / apply device tags  R1 / R2
  3. Associate devices with Configuration Group
  4. Upload WAN variables and deploy Configuration Group
  5. Generate and download Bootstrap (cloud-init) configs
"""

import sys
import csv
import json
import time
import getpass
import urllib3
import requests
import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration – edit these if the environment changes
# ---------------------------------------------------------------------------
VMANAGE_URL       = "https://vmanage-953677893.sdwan.cisco.com"
CONFIG_GROUP_NAME = "onboard_r1_pppoe_r2_pppoe"
CSV_PATH = (
    r"/mnt/c/Users/nick.oneill/OneDrive - Maintel Europe Limited"
    r"/Southern Coops - Rollout docs/vmanage-import-sc.csv"
)
ROLLOUT_TRACKER_PATH = (
    r"/mnt/c/Users/nick.oneill/OneDrive - Maintel Europe Limited"
    r"/Southern Coops - Rollout docs/NOF2025 Rollout tracker.xlsx"
)
SCOOP_INVENTORY_PATH = (
    r"/mnt/c/Users/nick.oneill/OneDrive - Maintel Europe Limited"
    r"/Southern Coops - IP Addresses/SCOOP Remote Site Inventory.xlsx"
)

# Suppress SSL warnings for self-signed controller certificate
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ---------------------------------------------------------------------------
# Console helpers
# ---------------------------------------------------------------------------
SEP = "=" * 70


def banner(msg: str) -> None:
    print(f"\n{SEP}\n  {msg}\n{SEP}")


def ok(msg: str)   -> None: print(f"  [OK]   {msg}")
def info(msg: str) -> None: print(f"  [..]   {msg}")
def warn(msg: str) -> None: print(f"  [WARN] {msg}")


def abort(msg: str, hint: str = "") -> None:
    print(f"\n  [FAIL] {msg}")
    if hint:
        print(f"  [HINT] {hint}")
    print()
    sys.exit(1)


# ---------------------------------------------------------------------------
# vManage session
# ---------------------------------------------------------------------------
class VManageSession:
    """Authenticated requests session for vManage 20.15.x."""

    def __init__(self, base_url: str, username: str, password: str):
        self.base_url = base_url.rstrip("/")
        self.session  = requests.Session()
        self.session.verify = False
        self._login(username, password)
        self._fetch_xsrf_token()

    def _login(self, username: str, password: str) -> None:
        resp = self.session.post(
            f"{self.base_url}/j_security_check",
            data={"j_username": username, "j_password": password},
            timeout=30,
            allow_redirects=True,
        )
        # vManage returns HTTP 200 with an HTML error page on bad credentials
        if (resp.status_code != 200
                or "Invalid username" in resp.text
                or "loginform" in resp.text.lower()):
            abort(
                "Authentication failed – bad username or password.",
                "Also confirm your IP is not blocked in vManage "
                "(Administration > Settings > Allowed IPs).",
            )
        ok("Authenticated with vManage.")

    def _fetch_xsrf_token(self) -> None:
        """vManage 20.x requires an X-XSRF-TOKEN header for write operations."""
        resp = self.session.get(f"{self.base_url}/dataservice/client/token", timeout=15)
        if resp.status_code == 200 and resp.text.strip():
            self.session.headers.update({"X-XSRF-TOKEN": resp.text.strip()})
            ok("XSRF token obtained.")
        else:
            warn("Could not retrieve XSRF token – write operations may fail.")

    # HTTP convenience wrappers ----------------------------------------------
    def get(self, path, **kw):
        kw.setdefault("timeout", 60)
        return self.session.get(f"{self.base_url}{path}", **kw)

    def post(self, path, **kw):
        kw.setdefault("timeout", 60)
        return self.session.post(f"{self.base_url}{path}", **kw)

    def put(self, path, **kw):
        kw.setdefault("timeout", 60)
        return self.session.put(f"{self.base_url}{path}", **kw)


# ---------------------------------------------------------------------------
# Step 0 – Load CSV data for the requested site
# ---------------------------------------------------------------------------
def load_site_data(site_id: str) -> list:
    """
    Return CSV rows where column D (index 3) == site_id.
    Also extracts the serial number from the Device ID (column A).

    Device ID format:  MODEL-VARIANT-SERIALNUMBER
    e.g.  C1121X-8P-FCZ2937R0T3  ->  serial = FCZ2937R0T3
    """
    csv_file = Path(CSV_PATH)
    if not csv_file.exists():
        abort(
            f"CSV file not found:\n  {CSV_PATH}",
            "Check that OneDrive is synced and the path above is correct.",
        )

    rows = []
    with open(csv_file, newline="", encoding="utf-8-sig") as fh:
        reader  = csv.DictReader(fh)
        headers = reader.fieldnames
        if not headers:
            abort("CSV has no header row – cannot parse.")

        info(f"CSV has {len(headers)} columns. First few: {list(headers[:6])}")

        for row in reader:
            values   = list(row.values())
            if len(values) < 4:
                continue
            row_site = values[3].strip()
            if row_site != site_id.strip():
                continue

            # Normalise: strip whitespace from all keys and values
            clean = {k.strip(): v.strip() for k, v in zip(headers, values)}

            # Extract serial number from Device ID (column A)
            device_id = clean.get("Device ID", "")
            if not device_id:
                abort(
                    "A CSV row for this site has an empty 'Device ID' (column A).",
                    f"Row data: {clean}",
                )
            # Serial = everything after the last '-'
            serial = device_id.rsplit("-", 1)[-1].upper()
            clean["_serial"]          = serial   # internal use only
            clean["_device_id_full"]  = device_id

            rows.append(clean)

    return rows


# ---------------------------------------------------------------------------
# Step 1 – Fetch WAN edge inventory; match devices by serial number
# ---------------------------------------------------------------------------
def get_wan_edge_inventory(vm: VManageSession) -> list:
    resp = vm.get("/dataservice/system/device/vedges")
    if resp.status_code != 200:
        abort(
            f"Failed to retrieve WAN edge inventory (HTTP {resp.status_code}).",
            f"Response: {resp.text[:400]}",
        )
    return resp.json().get("data", [])


def match_devices_by_uuid(edges: list, csv_rows: list) -> list:
    """
    Match vManage WAN edge records to CSV rows using the full Device ID
    from CSV column A (e.g. 'C1121X-8P-FCZ2937R0T3').

    vManage stores this value in the 'uuid' field of the vedge inventory.
    The serial number (suffix after last '-') is kept as a fallback and for
    display purposes only.

    Lookup order per device:
      1. Exact match on 'uuid'  (case-insensitive)
      2. Exact match on 'chasisNumber' / 'chassisNumber'
      3. Suffix match on the serial portion (last resort, with warning)
    """
    # Build lookup maps keyed on uppercased field values
    uuid_map    = {}   # full Device ID  -> edge
    serial_map  = {}   # serial suffix   -> edge

    for edge in edges:
        # Primary: uuid field (full device ID like C1121X-8P-FCZ2937R0T3)
        for field in ("uuid", "chasisNumber", "chassisNumber"):
            val = edge.get(field, "").strip().upper()
            if val:
                uuid_map[val] = edge
                break

        # Secondary: serialNumber / numeric licence serial (for fallback only)
        sn = edge.get("serialNumber", "").strip().upper()
        if sn:
            serial_map[sn] = edge

    devices    = []
    tag_labels = ["R1", "R2"]

    for idx, row in enumerate(csv_rows[:2]):
        device_id = row["_device_id_full"].strip().upper()   # e.g. C1121X-8P-FCZ2937R0T3
        serial    = row["_serial"].upper()                   # e.g. FCZ2937R0T3
        tag       = tag_labels[idx]
        matched   = None
        match_how = ""

        # 1. Exact full Device ID match
        if device_id in uuid_map:
            matched   = uuid_map[device_id]
            match_how = f"exact uuid match ('{device_id}')"

        # 2. Exact serial suffix match against uuid_map keys
        if not matched:
            for k, v in uuid_map.items():
                if k.endswith(serial):
                    matched   = v
                    match_how = f"suffix match on uuid key '{k}'"
                    warn(f"Full Device ID not found; using suffix match: {k}")
                    break

        # 3. Last resort: serialNumber field
        if not matched and serial in serial_map:
            matched   = serial_map[serial]
            match_how = f"serialNumber field match ('{serial}')"
            warn(f"Fell back to serialNumber field match for {tag}.")

        if not matched:
            sample_uuids = list(uuid_map.keys())[:20]
            abort(
                f"Cannot find {tag} in vManage WAN edge inventory.",
                f"CSV Device ID (column A) : {row['_device_id_full']}\n"
                f"  Looked up as (upper)   : {device_id}\n"
                f"  Serial suffix          : {serial}\n"
                f"  Sample vManage UUIDs (first 20): {sample_uuids}\n"
                "  Check Configuration > WAN Edges in vManage and confirm\n"
                "  the Device ID in column A of the CSV matches exactly.",
            )

        device = dict(matched)        # copy to avoid mutating the inventory
        device["_tag"]     = tag
        device["_serial"]  = serial
        device["_csv_row"] = row

        # Ensure a consistent 'uuid' key for all downstream steps
        if not device.get("uuid"):
            device["uuid"] = (
                device.get("chasisNumber")
                or device.get("chassisNumber")
                or device_id
            )

        devices.append(device)
        ok(
            f"{tag} matched ({match_how})\n"
            f"       uuid     = {device.get('uuid','?')}\n"
            f"       hostname = {device.get('host-name', device.get('hostName','(not set yet)'))}\n"
            f"       state    = {device.get('reachability', device.get('vedgeCSRUploadStatus','?'))}"
        )

    return devices


# ---------------------------------------------------------------------------
# Step 2 – Manual device tag confirmation (R1 / R2)
# ---------------------------------------------------------------------------

def check_and_apply_tags(vm: VManageSession, devices: list) -> None:
    banner("STEP 2 – Apply Device Tags (R1 / R2)  [MANUAL STEP]")

    print()
    print("  Please apply the following tags in vManage before continuing:")
    print()
    print("  Navigation: Configuration > WAN Edges")
    print("  For each device below, locate it by UUID, click the [...] menu,")
    print("  select 'Edit', and set the Tag field to the value shown.")
    print()
    print(f"  {'Tag':<6}  {'UUID':<30}  {'Serial'}")
    print(f"  {'-'*6}  {'-'*30}  {'-'*20}")
    for device in devices:
        tag    = device["_tag"]
        uuid   = device.get("uuid", "")
        serial = device["_serial"]
        print(f"  {tag:<6}  {uuid:<30}  {serial}")
    print()

    while True:
        answer = input("  Have the tags been applied? [yes/no/skip]: ").strip().lower()
        if answer in ("yes", "y"):
            ok("User confirmed tags applied – continuing.")
            break
        elif answer in ("skip", "s"):
            warn("Tag step skipped – ensure tags are correct before deployment.")
            break
        elif answer in ("no", "n"):
            print("  Please apply the tags in vManage and then type 'yes' to continue,")
            print("  or 'skip' to proceed without tags (not recommended).")
        else:
            print("  Please type 'yes', 'no', or 'skip'.")


# ---------------------------------------------------------------------------
# Step 3 – Associate devices with Configuration Group
# ---------------------------------------------------------------------------
def get_config_group(vm: VManageSession, group_name: str) -> dict:
    # Older vManage uses /dataservice/v1/config-group but some variants might differ. Let's try the v1 endpoint first.
    resp = vm.get("/dataservice/v1/config-group")
    if resp.status_code != 200:
        # Fallback for newer or different versions
        fallback_resp = vm.get("/dataservice/configurationGroups")
        if fallback_resp.status_code != 200:
            abort(
                f"Could not list Configuration Groups (HTTP {resp.status_code}).",
                f"Response: {resp.text[:400]}\nFallback Response: {fallback_resp.text[:400]}",
            )
        resp = fallback_resp

    raw = resp.json()
    groups = raw if isinstance(raw, list) else raw.get("data", [])

    for g in groups:
        if g.get("name", "") == group_name:
            return g

    names = [g.get("name") for g in groups]
    abort(
        f"Configuration Group '{group_name}' not found in vManage.",
        f"Groups found: {names}\n"
        "Verify the name under Configuration > Configuration Groups.",
    )


def associate_devices(vm: VManageSession, group_id: str, devices: list) -> None:
    banner("STEP 3 – Associate Devices with Configuration Group")

    uuids = [d["uuid"] for d in devices if d.get("uuid")]
    if not uuids:
        abort("No valid device UUIDs available for association.")

    # In newer vManage (like 20.15.x which uses /v1 endpoints), the associate endpoint is often
    # PUT or POST to /dataservice/v1/config-group/{group_id}/device/associate
    
    # Try the v1 associate endpoint first
    resp = vm.post(
        f"/dataservice/v1/config-group/{group_id}/device/associate",
        json={"devices": [{"id": u} for u in uuids]},
    )
    
    # Fallback to the older format if the 404 persists
    if resp.status_code == 404:
        resp = vm.post(
            f"/dataservice/configurationGroups/{group_id}/devices",
            json={"deviceIds": uuids},
        )

    if resp.status_code in (200, 201, 202, 204):
        ok(f"Devices associated with Configuration Group (id={group_id}).")
    elif resp.status_code == 409 or (resp.status_code == 400 and "CFGRP0018" in resp.text):
        ok("Devices already associated with the Configuration Group.")
    else:
        abort(
            f"Association failed (HTTP {resp.status_code}): {resp.text}",
            "Check that the Configuration Group exists and your user has write access.",
        )

    for d in devices:
        info(f"  {d['_tag']}  uuid={d['uuid']}  serial={d['_serial']}")


# ---------------------------------------------------------------------------
# Step 4 – Upload WAN variables and deploy
# ---------------------------------------------------------------------------
# Keys that are script-internal and must NOT be sent to vManage
_INTERNAL_KEYS = {"_serial", "_tag", "_csv_row", "_device_id_full"}

# CSV column names that map to vManage variable names
_CSV_HEADER_MAP = {
    "System IP":  "system_ip",
    "Host Name":  "host_name",
    "Site Id":    "site_id",
}

# Variable names whose values must be sent as numbers (int or float)
_NUMERIC_VARS = {
    "site_id", "basic_gpsl_latitude", "basic_gpsl_longitude",
    "tloc_bandwidth_up", "tloc_bandwidth_down",
    "wan_bandwidth_up", "wan_bandwidth_down",
    "wan_shapingRate", "ethpppoe_ipsecPrefer",
    "Rollback Timer (sec)",
}

# CSV columns NOT defined in the onboard_r1_pppoe_r2_pppoe config group schema.
# vManage rejects any variable that is not declared in the template.
_VARS_EXCLUDED = {
    "Dual Stack IPv6 Default",
    "Rollback Timer (sec)",
    "cloudSaaSDeviceRole_variable",
    "cloudSaaSVpnType_variable",
    "cloudSaasLBEnabled_variable",
    "cloudSaasLatency_variable",
    "cloudSaasLoss_variable",
    "cloudSaasSigEnabled_variable",
    "cloudSaasSourceIpBased_variable",
    "cloudSaasTlocList_variable",
    "lan_vpn_100_nat_1_rangeEnd",
    "lan_vpn_100_nat_1_rangeStart",
    "lan_vpn_100_staticNat_1_translatedSourceIp",
    "lan_vpn_100_staticNat_2_translatedSourceIp",
    "port_offset",
    "provision_port_disable",
    "qos_Interface_1",
    "static_wan_gw",
    "static_wan_ip",
    "static_wan_mask",
    "vlan100_dhcp_exclude", "vlan100_ipv4", "vlan100_mask", "vlan100_vrrp_ipv4", "vlan100_vrrp_pri",
    "vlan101_dhcp_exclude", "vlan101_dhcp_gateway", "vlan101_dhcp_mask", "vlan101_dhcp_net",
    "vlan101_ipv4", "vlan101_mask", "vlan101_vrrp_ipv4", "vlan101_vrrp_pri",
    "vlan10_dhcp_exclude", "vlan10_dhcp_gateway", "vlan10_dhcp_mask", "vlan10_dhcp_net",
    "vlan10_ipv4", "vlan10_mask", "vlan10_vrrp_ipv4", "vlan10_vrrp_pri",
    "vlan120_dhcp_exclude", "vlan120_ipv4", "vlan120_mask", "vlan120_vrrp_ipv4", "vlan120_vrrp_pri",
    "vlan20_dhcp_exclude", "vlan20_dhcp_gateway", "vlan20_dhcp_mask", "vlan20_dhcp_net",
    "vlan20_ipv4", "vlan20_mask", "vlan20_vrrp_ipv4", "vlan20_vrrp_pri",
    "vlan2_dhcp_exclude", "vlan2_dhcp_gateway", "vlan2_dhcp_mask", "vlan2_dhcp_net",
    "vlan2_ipv4", "vlan2_mask", "vlan2_vrrp_ipv4", "vlan2_vrrp_pri",
    "vlan30_dhcp_exclude", "vlan30_ipv4", "vlan30_mask", "vlan30_vrrp_ipv4", "vlan30_vrrp_pri",
    "vlan31_dhcp_exclude", "vlan31_dhcp_gateway", "vlan31_dhcp_mask", "vlan31_dhcp_net",
    "vlan31_ipv4", "vlan31_mask", "vlan31_vrrp_ipv4", "vlan31_vrrp_pri",
    "vlan40_dhcp_exclude", "vlan40_ipv4", "vlan40_mask", "vlan40_vrrp_ipv4", "vlan40_vrrp_pri",
    "vlan60_ipv4", "vlan60_mask", "vlan60_vrrp_ipv4", "vlan60_vrrp_pri",
    "vlan70_ipv4", "vlan70_mask", "vlan70_vrrp_ipv4", "vlan70_vrrp_pri",
    "vlan80_ipv4", "vlan80_mask", "vlan80_vrrp_ipv4", "vlan80_vrrp_pri",
    "wan_track_addr_tloc",
}


def _coerce(name: str, value: str):
    """Cast value to the correct type for vManage."""
    if name in _NUMERIC_VARS:
        try:
            f = float(value)
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            pass
    return value


def _build_variables(csv_row: dict) -> list:
    """
    Return variables as a list of {name, value} dicts for the v1 vManage API.
    Excludes CSV columns not defined in the config group schema.
    """
    result = []
    for k, v in csv_row.items():
        if k in _INTERNAL_KEYS or k == "Device ID" or k.startswith("_") or v == "":
            continue
        name = _CSV_HEADER_MAP.get(k, k)
        if name in _VARS_EXCLUDED or k in _VARS_EXCLUDED:
            continue
        result.append({"name": name, "value": _coerce(name, v)})
    return result


def deploy_config_group(vm: VManageSession, group_id: str, devices: list) -> None:
    banner("STEP 4 – Upload WAN Variables and Deploy Configuration Group")

    device_var_list = []

    for device in devices:
        variables = _build_variables(device["_csv_row"])

        # pseudo_commit_timer is required by the schema but not present in the CSV
        if not any(v["name"] == "pseudo_commit_timer" for v in variables):
            variables.append({"name": "pseudo_commit_timer", "value": 0})

        device_var_list.append({
            "device-id": device["uuid"],
            "variables": variables,
        })
        info(
            f"  {device['_tag']} ({device['uuid']}): "
            f"{len(variables)} variables prepared."
        )

    # Upload variables
    info("Uploading WAN variables to vManage...")
    var_resp = vm.put(
        f"/dataservice/v1/config-group/{group_id}/device/variables",
        json={"solution": "sdwan", "devices": device_var_list},
    )

    if var_resp.status_code == 404:
        var_resp = vm.post(
            f"/dataservice/v1/config-group/{group_id}/device/variables",
            json={"solution": "sdwan", "devices": device_var_list},
        )

    if var_resp.status_code == 404:
        var_resp = vm.post(
            f"/dataservice/configurationGroups/{group_id}/devices/variables",
            json={"solution": "sdwan", "device-list": device_var_list},
        )

    if var_resp.status_code not in (200, 201, 204):
        abort(
            f"Variable upload failed (HTTP {var_resp.status_code}): {var_resp.text}",
            "Ensure the CSV column names exactly match the variable names\n"
            "defined in the Configuration Group's feature templates.\n"
            "Check Configuration > Configuration Groups > <group> > Variables.",
        )
    ok("WAN variables uploaded successfully.")

    # Trigger deployment
    uuids       = [d["uuid"] for d in devices]
    
    # Try the v1 endpoints first
    dep_resp    = vm.post(
        f"/dataservice/v1/config-group/{group_id}/device/deploy",
        json={"devices": [{"id": u} for u in uuids]},
    )
    
    if dep_resp.status_code == 404:
        # Fallback to older ones
        dep_resp    = vm.post(
            f"/dataservice/configurationGroups/{group_id}/devices/deploy",
            json={"deviceIds": uuids},
        )
        
    if dep_resp.status_code not in (200, 201, 202):
        abort(
            f"Deployment request failed (HTTP {dep_resp.status_code}): {dep_resp.text}",
            "Check device reachability and Configuration Group validity in vManage.",
        )

    action_id = dep_resp.json().get("id", "")
    ok(f"Deployment triggered (action id: {action_id or 'not returned'}).")

    if not action_id:
        info("No action ID in response – querying recent tasks to find the deploy job...")
        action_id = _find_recent_deploy_action(vm, uuids)

    if action_id:
        _poll_action(vm, action_id, label="Deployment", timeout=300)
    else:
        warn(
            "Could not locate a task ID for this deployment.\n"
            "  Monitor the job at: Monitor > Maintenance > Tasks in vManage."
        )
        print()
        print("  Please check Monitor > Maintenance > Tasks and confirm the")
        print("  deployment has completed (or is in progress) before continuing.")
        print()
        ans = input("  Has the deployment completed? [y/N]: ").strip().lower()
        if ans not in ("y", "yes"):
            abort(
                "Deployment not confirmed by user.",
                "Re-run the script once the deployment task has completed in vManage.",
            )


def _find_recent_deploy_action(vm: VManageSession, uuids: list) -> str:
    """
    Poll /dataservice/device/action/status/tasks for a running deploy job
    that involves any of the given device UUIDs. Retries for up to 30 s
    to allow vManage time to register the task after the deploy call.
    Returns the processId string, or "" if nothing matched.
    """
    uuid_set = set(u.lower() for u in uuids)
    deadline = time.time() + 30

    while time.time() < deadline:
        resp = vm.get("/dataservice/device/action/status/tasks")
        if resp.status_code != 200:
            time.sleep(5)
            continue

        try:
            tasks = resp.json().get("runningTasks", [])
            for task in tasks:
                process_id = str(task.get("processId", ""))
                if not process_id:
                    continue

                # Match on device UUIDs embedded in the task
                task_devices = (
                    task.get("deviceUUIDs", [])
                    or task.get("deviceIds", [])
                    or [d.get("deviceId", "") for d in task.get("deviceList", [])]
                )
                task_devices_lower = [str(d).lower() for d in task_devices]

                # Also match on action type
                action_type = " ".join([
                    str(task.get("actionConfig", "")),
                    str(task.get("name", "")),
                    str(task.get("action", "")),
                    str(task.get("type", "")),
                ]).lower()

                device_match = any(u in task_devices_lower for u in uuid_set)
                type_match   = "deploy" in action_type or "config" in action_type

                if device_match or (type_match and tasks):
                    info(f"  Matched deploy task: processId={process_id}")
                    return process_id
        except Exception:
            pass

        time.sleep(5)

    return ""


def _poll_action(vm: VManageSession, action_id: str,
                 label: str = "Action", timeout: int = 300) -> None:
    info(f"Polling {label} status every 15 s (timeout={timeout} s)...")
    deadline = time.time() + timeout

    while time.time() < deadline:
        # Primary: per-action status endpoint
        resp = vm.get(f"/dataservice/device/action/status/{action_id}")

        if resp.status_code == 404:
            # Fallback: check runningTasks list and look for our processId
            tr = vm.get("/dataservice/device/action/status/tasks")
            if tr.status_code == 200:
                running = tr.json().get("runningTasks", [])
                match = next((t for t in running if str(t.get("processId", "")) == action_id), None)
                if match is None:
                    # No longer in running list – may have completed
                    ok(f"{label} is no longer in the running task list – assumed complete.")
                    return
                status = str(match.get("status", "in_progress")).lower()
                info(f"  {label}: {status} (from runningTasks)")
                if status in ("success", "done", "success_scheduled"):
                    ok(f"{label} completed successfully.")
                    return
                time.sleep(15)
                continue

        if resp.status_code != 200:
            warn(f"Status poll returned HTTP {resp.status_code} – retrying...")
            time.sleep(10)
            continue

        data    = resp.json()
        summary = data.get("summary", {})
        status  = summary.get("status", "").lower()
        counts  = summary.get("count", {})
        info(
            f"  {label}: {status}  |  "
            f"Success={counts.get('Success','?')}  "
            f"Failure={counts.get('Failure','?')}"
        )

        if status in ("success", "done", "success_scheduled"):
            ok(f"{label} completed successfully.")
            return

        if status in ("failure", "failed", "error"):
            details = json.dumps(data.get("data", [])[:5], indent=2)
            abort(
                f"{label} reported failure (status={status}).",
                f"First entries from vManage:\n{details}\n"
                "Full details: Monitor > Maintenance > Tasks in vManage.",
            )

        time.sleep(15)

    abort(
        f"{label} timed out after {timeout} s.",
        "The job may still be running. Check Monitor > Maintenance > Tasks.",
    )


# ---------------------------------------------------------------------------
# Step 5 – Generate and download Bootstrap (cloud-init) configs
# ---------------------------------------------------------------------------
def download_bootstrap_configs(vm: VManageSession, devices: list, site_id: str) -> None:
    banner("STEP 5 – Generate & Download Bootstrap (Cloud-Init) Configs")

    output_dir = Path("/mnt/c/Users/nick.oneill/OneDrive - Maintel Europe Limited/Southern Coops - Rollout docs/bootstrap-configs")
    output_dir.mkdir(parents=True, exist_ok=True)
    info(f"Output directory: {output_dir}")

    for device in devices:
        tag      = device["_tag"]
        uuid     = device.get("uuid", "")
        hostname = device.get("host-name", device.get("hostName", device["_serial"]))

        info(f"Requesting bootstrap for {tag} ({hostname}, uuid={uuid})...")

        # Bootstrap generation can be slow — use a longer timeout
        _bootstrap_timeout = 180

        # Try endpoints in order; fall through on 404 or 405 (wrong method/path)
        _bootstrap_resp = None

        _bootstrap_attempts = [
            # vManage 20.15: GET /device/bootstrap/device/{uuid}?configtype=cloudinit
            lambda: vm.get(
                f"/dataservice/system/device/bootstrap/device/{uuid}",
                params={"configtype": "cloudinit"},
                timeout=_bootstrap_timeout,
            ),
            # vManage 20.12: POST /device/bootstrap
            lambda: vm.post(
                "/dataservice/system/device/bootstrap",
                json={"deviceId": uuid, "configType": "cloudinit", "generateHasVpnInterface": True},
                timeout=_bootstrap_timeout,
            ),
            # Legacy endpoint
            lambda: vm.post(
                "/dataservice/bootstrapConfig",
                json={"deviceId": uuid, "configType": "cloudinit", "generateHasVpnInterface": True},
                timeout=_bootstrap_timeout,
            ),
        ]

        for attempt in _bootstrap_attempts:
            try:
                _bootstrap_resp = attempt()
                if _bootstrap_resp.status_code in (200, 201):
                    break
                if _bootstrap_resp.status_code not in (404, 405):
                    # A real error — stop trying
                    break
                warn(f"  Endpoint returned HTTP {_bootstrap_resp.status_code}, trying next...")
            except Exception as exc:
                warn(f"  Bootstrap attempt timed out ({exc}), trying next...")
                _bootstrap_resp = None

        resp = _bootstrap_resp

        if resp is None:
            abort(
                f"Bootstrap request for {tag} timed out on all endpoints.",
                "vManage may be under load. Try re-running the script later.",
            )

        if resp.status_code not in (200, 201):
            abort(
                f"Bootstrap generation failed for {tag} / {hostname} "
                f"(HTTP {resp.status_code}): {resp.text}",
                "Requirements for bootstrap generation:\n"
                "  - Device must have a valid certificate installed\n"
                "  - Device must be in 'certificate installed' state in\n"
                "    Configuration > WAN Edges",
            )

        data = resp.json()

        # vManage may use any of these key names
        config_text = (
            data.get("bootstrapConfig")
            or data.get("config")
            or data.get("bootstrapconfig")
            or json.dumps(data, indent=2)   # fallback: save raw JSON
        )

        filename = output_dir / f"{hostname}-{device['_serial']}.cfg"
        filename.write_text(config_text, encoding="utf-8")
        ok(f"Saved: {filename}")

    ok(f"All bootstrap configs saved to: {output_dir}/")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def _onboard_site(vm, username: str, password: str, site_id: str) -> None:
    """Run the full onboarding workflow for a single site."""
    # ------------------------------------------------------------------
    # Step 0 – Load CSV
    # ------------------------------------------------------------------
    banner(f"STEP 0 – Loading CSV Data  (site_id = {site_id})")
    info(f"CSV: {CSV_PATH}")

    csv_rows = load_site_data(site_id)
    if not csv_rows:
        abort(
            f"No rows found for site_id '{site_id}' in the CSV.",
            "Check that column D (Site Id) contains this value and the file is accessible.",
        )
    ok(f"Found {len(csv_rows)} row(s) for site {site_id}.")
    for i, row in enumerate(csv_rows[:2]):
        tag = "R1" if i == 0 else "R2"
        info(f"  {tag}: Device ID = {row['_device_id_full']}  ->  Serial = {row['_serial']}")

    if len(csv_rows) > 2:
        warn(f"CSV has {len(csv_rows)} rows for this site – only the first 2 (R1, R2) will be used.")

    # ------------------------------------------------------------------
    # Step 1 – Connect + locate devices
    # ------------------------------------------------------------------
    banner("STEP 1 – Connecting to vManage & Locating Devices")

    info("Fetching WAN edge inventory...")
    all_edges = get_wan_edge_inventory(vm)
    info(f"Inventory contains {len(all_edges)} WAN edge record(s).")

    devices = match_devices_by_uuid(all_edges, csv_rows)

    target_serials = sorted(d["_serial"] for d in devices)
    ok(f"Scope confirmed – will only operate on serials: {target_serials}")

    # ------------------------------------------------------------------
    # Step 2 – Tags
    # ------------------------------------------------------------------
    check_and_apply_tags(vm, devices)

    # ------------------------------------------------------------------
    # Step 3 – Configuration Group association
    # ------------------------------------------------------------------
    banner("Locating Configuration Group")
    cfg_group = get_config_group(vm, CONFIG_GROUP_NAME)
    group_id  = cfg_group.get("id") or cfg_group.get("configGroupId", "")
    if not group_id:
        abort(
            f"Configuration Group '{CONFIG_GROUP_NAME}' returned no usable ID.",
            f"Raw response: {json.dumps(cfg_group)[:500]}",
        )
    ok(f"Found '{CONFIG_GROUP_NAME}' (id={group_id}).")

    associate_devices(vm, group_id, devices)

    # ------------------------------------------------------------------
    # Step 4 – Deploy
    # ------------------------------------------------------------------
    deploy_config_group(vm, group_id, devices)

    # ------------------------------------------------------------------
    # Step 5 – Bootstrap
    # ------------------------------------------------------------------
    download_bootstrap_configs(vm, devices, site_id)

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    banner("ALL TASKS COMPLETE")
    ok(f"Site ID      : {site_id}")
    for d in devices:
        ok(
            f"{d['_tag']}          : serial={d['_serial']}  "
            f"uuid={d.get('uuid','?')}  "
            f"host={d.get('host-name', d.get('hostName','?'))}"
        )
    ok(f"Config Group : {CONFIG_GROUP_NAME}")
    ok(f"Bootstrap    : /mnt/c/Users/nick.oneill/OneDrive - Maintel Europe Limited/Southern Coops - Rollout docs/bootstrap-configs/")
    print()
    print("  Next steps:")
    print("  1. Copy the cloud-init .cfg file to a USB stick or SD card")
    print("  2. Insert into the WAN edge and power on – ZTP will auto-onboard")
    print("  3. Monitor progress: Monitor > Devices in vManage")
    print()


def generate_ping_template(site_ids: list) -> None:
    """
    Look up network details for each onboarded site and print a PING IP TEMPLATE.
    site_ids: list of SD-WAN site IDs entered during the session (e.g. '30409').
    """
    if not site_ids:
        return

    # ------------------------------------------------------------------
    # Load rollout tracker (header is row index 1, data starts at row 2)
    # col indices (0-based): A=0, G=6, N=13, Q=16, X=23
    # ------------------------------------------------------------------
    try:
        rt_wb = openpyxl.load_workbook(ROLLOUT_TRACKER_PATH, read_only=True, data_only=True)
        rt_ws = rt_wb["Main"]
    except Exception as exc:
        warn(f"Could not open rollout tracker: {exc}")
        rt_ws = None

    # Build store_number -> row dict (strip leading zeros from col A)
    store_rows = {}
    if rt_ws:
        for row in rt_ws.iter_rows(min_row=3, values_only=True):
            if row[0] is not None:
                store_num = str(row[0]).strip().strip("'").lstrip("0") or "0"
                # If we already have a populated entry for this store, keep it
                if store_num in store_rows and any(store_rows[store_num][i] for i in (6, 13, 16, 23)):
                    continue
                store_rows[store_num] = row
        rt_wb.close()

    # ------------------------------------------------------------------
    # Load SCOOP inventory – Maintel ISP sheet
    # col C (idx 2) = TTB/PXC username  → WAN IP from col B (idx 1)
    # col G (idx 6) = BT username       → WAN IP from col F (idx 5)
    # ------------------------------------------------------------------
    wan_username_map: dict[str, str] = {}   # username -> WAN IP
    try:
        inv_wb = openpyxl.load_workbook(SCOOP_INVENTORY_PATH, read_only=True, data_only=True)
        inv_ws = inv_wb["Maintel ISP"]
        for row in inv_ws.iter_rows(min_row=3, values_only=True):
            # TTB/PXC: col C username (idx 2) → col B IP (idx 1)
            if row[2] and row[1]:
                wan_username_map[str(row[2]).strip().strip("'")] = str(row[1]).strip().strip("'")
            # BT: col G username (idx 6) → col F IP (idx 5)
            if len(row) > 6 and row[6] and row[5]:
                wan_username_map[str(row[6]).strip().strip("'")] = str(row[5]).strip().strip("'")
        inv_wb.close()
    except Exception as exc:
        warn(f"Could not open SCOOP inventory: {exc}")

    # ------------------------------------------------------------------
    # Build per-store info and print template
    # ------------------------------------------------------------------
    print()
    print("PING IP TEMPLATE")
    print("================")
    print()

    for site_id in site_ids:
        store_num = site_id[-4:].lstrip("0") or "0"

        row = store_rows.get(store_num)

        if not row:
            print(f"# (store {store_num} not found in rollout tracker)")
            continue

        r1_mgmt  = str(row[6]).strip().strip("'")  if row[6]  else ""
        r1_ppp   = str(row[13]).strip().strip("'") if row[13] else ""
        r2_mgmt  = str(row[16]).strip().strip("'") if row[16] else ""
        r2_ppp   = str(row[23]).strip().strip("'") if row[23] else ""

        r1_wan = wan_username_map.get(r1_ppp, "") if r1_ppp else ""
        r2_wan = wan_username_map.get(r2_ppp, "") if r2_ppp else ""

        print("#")
        print(f"{store_num}")
        print("#")
        if r1_wan:
            print(r1_wan)
        if r1_mgmt:
            print(r1_mgmt)
        if r2_wan:
            print(r2_wan)
        if r2_mgmt:
            print(r2_mgmt)
        if not any([r1_wan, r1_mgmt, r2_wan, r2_mgmt]):
            print(f"# (no IP data in tracker for store {store_num})")


def main() -> None:
    print()
    print(SEP)
    print("  Cisco SD-WAN – vManage WAN Edge Deployment Script")
    print(f"  Controller : {VMANAGE_URL}")
    print(f"  Code ver.  : 20.15.4.4")
    print(SEP)

    print()
    username = input("vManage Username : ").strip()
    if not username:
        abort("Username cannot be empty.")
    password = getpass.getpass("vManage Password : ")
    if not password:
        abort("Password cannot be empty.")

    # Authenticate once; reuse the session for all sites
    banner("Connecting to vManage")
    try:
        vm = VManageSession(VMANAGE_URL, username, password)
    except requests.exceptions.ConnectionError as exc:
        abort(
            f"Cannot reach vManage at {VMANAGE_URL}.",
            f"Check VPN / network access.\nError: {exc}",
        )
    ok("Authenticated successfully.")

    completed_sites: list[str] = []

    while True:
        print()
        site_id = input("SD-WAN Site ID (or 'x' to exit) : ").strip()
        if site_id.lower() in ("x", "exit", "quit", "q"):
            generate_ping_template(completed_sites)
            print()
            print("  Exiting. Goodbye.")
            print()
            break
        if not site_id:
            warn("Site ID cannot be empty – try again.")
            continue
        try:
            _onboard_site(vm, username, password, site_id)
            completed_sites.append(site_id)
        except SystemExit:
            # abort() calls sys.exit(1); catch it so the loop continues
            print()
            ans = input("  Press Enter to onboard another site, or 'x' to exit: ").strip().lower()
            if ans in ("x", "exit", "quit", "q"):
                break


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  [ABORTED] User cancelled (Ctrl-C).")
        sys.exit(1)